# -*- coding: utf-8 -*-
from email import encoders
from email.mime.base import MIMEBase
from optparse import Values
import os
from os import link
from typing import List
import pandas as pd
import openpyxl
from openpyxl import workbook, load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from inspect import classify_class_attrs

#Lê o arquivo de texto para buscar o nome da planilha

#Ler arquivo de texto
f = open('config.txt','r')

#Armazena o conteudo escrito
txt=f.read()


#Carrega o arquivo Excel 
arquivo = load_workbook(txt+".xlsx")

abas = []

#Insere as abas da planilha no array abas
for aba in arquivo:
    abas.append(aba)

#Indica a aba ativa, no caso a aba 'BASE'
aba_ativa = arquivo['BASE']
medicos = {}

# Dicionario de medicos 
for coluna in aba_ativa['A']:
    if coluna.value != None:
        linha = coluna.row
        medicos[coluna.value] = aba_ativa[f"B{linha}"].value
        
medicos.pop('COD')

#Indica a aba ativa, no caso a aba 'Em branco'
aba_em_branco = arquivo['EM BRANCO']

#O for abaixo percorre o Dicionario de medicos, criando uma planilha individual para cada medico e após, salvando tudo 
for medico in medicos:    
    cell = aba_em_branco.cell(row = 3, column = 9).value = medico   

    arquivo.save(filename = medicos[medico] + '.xlsx')
    

